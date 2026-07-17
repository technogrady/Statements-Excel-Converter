"""End-to-end: CLI over the fixture folder → workbook verified with openpyxl."""
from datetime import datetime

import openpyxl
import pytest

import parse_statements
from excel_writer import sheet_name_for


@pytest.fixture(scope="module")
def workbook_path(fixture_dir, tmp_path_factory):
    out = tmp_path_factory.mktemp("out") / "Bank_Statements.xlsx"
    rc = parse_statements.main([str(fixture_dir), "-o", str(out)])
    assert rc == 0
    return out


@pytest.fixture(scope="module")
def wb(workbook_path):
    return openpyxl.load_workbook(workbook_path)


def inventory_rows(wb):
    ws = wb["Inventory"]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    out = []
    for row in rows[1:]:
        if row[0] is None or row[0] == "COVERAGE SUMMARY — one row per account":
            break
        out.append(dict(zip(header, row)))
    return out


class TestWorkbookStructure:
    def test_inventory_is_first_tab(self, wb):
        assert wb.sheetnames[0] == "Inventory"

    def test_one_sheet_per_account(self, wb):
        assert "Regions_x1000" in wb.sheetnames
        assert "ServisFirst_x5678" in wb.sheetnames

    def test_frozen_header_and_autofilter(self, wb):
        for name in ("Inventory", "Regions_x1000", "ServisFirst_x5678"):
            ws = wb[name]
            assert ws.freeze_panes == "A2", name
            assert ws.auto_filter.ref is not None, name


class TestInventory:
    def test_row_per_file_with_statuses(self, wb):
        rows = inventory_rows(wb)
        by_file = {r["File"]: r for r in rows}
        assert len(rows) == 8  # 5 statements (incl. dup) + 3 failures
        assert by_file["regions_checking_2022-01.pdf"]["Reconciled"] == "OK"
        assert by_file["regions_checking_2022-01_redownload.pdf"]["Reconciled"] == "DUPLICATE"
        assert "DUPLICATE of" in by_file["regions_checking_2022-01_redownload.pdf"]["Notes"]
        assert by_file["unknown_bank.pdf"]["Reconciled"] == "UNRECOGNIZED"
        assert "Example National Bank" in by_file["unknown_bank.pdf"]["Notes"]
        assert by_file["scanned_image_only.pdf"]["Reconciled"] == "NO_TEXT (possible scan)"
        assert by_file["password_protected.pdf"]["Reconciled"] == "ENCRYPTED"

    def test_statement_row_values(self, wb):
        rows = inventory_rows(wb)
        sf = next(r for r in rows if r["File"].startswith("servisfirst_checking"))
        assert sf["Bank"] == "ServisFirst"
        assert sf["Account"] == "x5678"
        assert sf["Account Type"] == "BUSINESS CHECKING"
        assert sf["# Transactions"] == 31
        assert round(sf["Opening Balance"], 2) == 20000.00
        assert round(sf["Closing Balance"], 2) == 26550.00
        assert round(sf["Total Credits"], 2) == 19000.00
        assert round(sf["Total Debits"], 2) == -12450.00

    def test_coverage_summary_block(self, wb):
        ws = wb["Inventory"]
        rows = list(ws.iter_rows(values_only=True))
        title_idx = next(i for i, r in enumerate(rows)
                         if r[0] == "COVERAGE SUMMARY — one row per account")
        header = rows[title_idx + 1]
        cov = [dict(zip(header, r)) for r in rows[title_idx + 2:] if r[0]]
        by_account = {r["Account"]: r for r in cov}
        regions = by_account["Regions_x1000"]
        assert regions["Balance Chain"] == "Gap(s)"
        assert regions["Missing Months"] == "2022-03"
        assert regions["Statements"] == 3
        assert "BUSINESS INTEREST CHECKING" in regions["Labels (period each used)"]
        assert "BUSINESS CHECKINGS" in regions["Labels (period each used)"]
        assert "Label changed (same account)" in regions["Notes"]
        assert by_account["ServisFirst_x5678"]["Balance Chain"] == "Single statement"
        assert by_account["ServisFirst_x5678"]["Missing Months"] == "(none)"


class TestAccountSheets:
    def test_sorted_by_date_ascending(self, wb):
        for name in ("Regions_x1000", "ServisFirst_x5678"):
            ws = wb[name]
            dates = [r[0] for r in ws.iter_rows(min_row=2, values_only=True) if r[0]]
            assert dates == sorted(dates), name

    def test_regions_transaction_count_excludes_duplicate(self, wb):
        ws = wb["Regions_x1000"]
        rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if r[0]]
        assert len(rows) == 11  # 7 + 3 + 1, duplicate statement excluded

    def test_columns(self, wb):
        ws = wb["ServisFirst_x5678"]
        header = [c.value for c in ws[1]]
        assert header == ["Date", "Type", "Check No", "Description", "Amount",
                          "Statement Period", "Source File"]

    def test_check_rows_carry_check_no(self, wb):
        ws = wb["ServisFirst_x5678"]
        checks = [r for r in ws.iter_rows(min_row=2, values_only=True)
                  if r[1] == "Check"]
        assert len(checks) == 13
        assert all(str(r[2]).startswith("50") for r in checks)
        assert all(r[4] < 0 for r in checks)

    def test_statement_period_and_source_populated(self, wb):
        ws = wb["Regions_x1000"]
        for r in ws.iter_rows(min_row=2, values_only=True):
            if r[0] is None:
                continue
            assert isinstance(r[0], datetime)
            assert " - " in r[5]
            assert r[6].endswith(".pdf")


class TestSheetNames:
    def test_forbidden_chars_and_length(self):
        taken = set()
        name = sheet_name_for("Bad[]:*?/\\Bank_x1234_with_very_long_tail", taken)
        assert len(name) <= 31
        assert not set("[]:*?/\\") & set(name)

    def test_uniqueness(self):
        taken = set()
        a = sheet_name_for("Regions_x1000", taken)
        b = sheet_name_for("Regions_x1000", taken)
        assert a != b
